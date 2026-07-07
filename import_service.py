import io
import re
import zipfile
import base64
import mimetypes
import os
import shutil
import subprocess
import tempfile
import uuid
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

from defusedxml.lxml import fromstring
from lxml import etree
from openpyxl import load_workbook
from PIL import Image, ImageChops


W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
M_NS = "http://schemas.openxmlformats.org/officeDocument/2006/math"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"

NS = {
	"w": W_NS,
	"m": M_NS,
	"r": R_NS,
	"a": A_NS,
}


def _env_truthy(name: str, default: bool = False) -> bool:
	val = os.getenv(name)
	if val is None:
		return default
	return str(val).strip().lower() in {"1", "true", "yes", "y", "on"}


def _escape_html(text: str) -> str:
	return (
		str(text)
		.replace("&", "&amp;")
		.replace("<", "&lt;")
		.replace(">", "&gt;")
		.replace('"', "&quot;")
		.replace("'", "&#039;")
	)


def _strip_tags(html: str) -> str:
	return re.sub(r"<[^>]+>", " ", str(html or "")).strip()


_SOFFICE_BIN = shutil.which("soffice") or shutil.which("libreoffice")

_RASTER_EXT_MIME = {
	"png": "image/png",
	"jpg": "image/jpeg",
	"jpeg": "image/jpeg",
	"gif": "image/gif",
	"bmp": "image/bmp",
	"tif": "image/tiff",
	"tiff": "image/tiff",
	"svg": "image/svg+xml",
}

# Word's legacy Equation Editor / pasted vector art (ChemDraw, shapes, etc.) is
# often embedded as EMF/WMF. Browsers cannot render either format directly, so
# a raw base64 data: URI just shows a broken image. LibreOffice (if installed
# on the host) can convert these to SVG, which every browser renders natively.
_VECTOR_EXT = {"emf", "wmf"}


def _autocrop_bbox(png_bytes: bytes, pad_frac: float = 0.04) -> Optional[Tuple[Tuple[int, int, int, int], Tuple[int, int]]]:
	"""Finds the pixel bounding box of non-white content in a page render, padded
	slightly, plus the full page's pixel size: ((x0, y0, x1, y1), (page_w, page_h)).
	None if detection fails or the page is blank (nothing to crop to)."""
	try:
		im = Image.open(io.BytesIO(png_bytes)).convert("RGB")
	except Exception:
		return None
	w, h = im.size
	if w <= 0 or h <= 0:
		return None
	bg = Image.new("RGB", im.size, (255, 255, 255))
	bbox = ImageChops.difference(im, bg).getbbox()
	if not bbox:
		return None
	x0, y0, x1, y1 = bbox
	pad_x, pad_y = (x1 - x0) * pad_frac, (y1 - y0) * pad_frac
	x0, y0 = max(0, x0 - pad_x), max(0, y0 - pad_y)
	x1, y1 = min(w, x1 + pad_x), min(h, y1 + pad_y)
	if x1 <= x0 or y1 <= y0:
		return None
	return (round(x0), round(y0), round(x1), round(y1)), (w, h)


def _crop_svg_viewbox(svg_bytes: bytes, frac_bbox: Tuple[float, float, float, float]) -> bytes:
	"""Narrows an SVG's viewBox to the given fractional sub-rectangle, without touching
	any path data - this crops out surrounding whitespace while staying fully vector."""
	try:
		text = svg_bytes.decode("utf-8")
	except Exception:
		return svg_bytes
	m = re.search(r'viewBox="([\d.\-]+)\s+([\d.\-]+)\s+([\d.\-]+)\s+([\d.\-]+)"', text)
	if not m:
		return svg_bytes
	vx, vy, vw, vh = (float(g) for g in m.groups())
	fx0, fy0, fx1, fy1 = frac_bbox
	new_x, new_y = vx + fx0 * vw, vy + fy0 * vh
	new_w, new_h = (fx1 - fx0) * vw, (fy1 - fy0) * vh
	if new_w <= 0 or new_h <= 0:
		return svg_bytes
	text = re.sub(
		r'viewBox="[^"]*"',
		f'viewBox="{new_x:.2f} {new_y:.2f} {new_w:.2f} {new_h:.2f}"',
		text,
		count=1,
	)
	# Drop the root <svg>'s physical width/height (e.g. "210mm"/"297mm" for the
	# original full page) - they no longer match the narrowed viewBox, and the
	# <img> tag we wrap this in already sets the final on-page size explicitly.
	text = re.sub(r'(<svg\b[^>]*?)\swidth="[^"]*"', r"\1", text, count=1)
	text = re.sub(r'(<svg\b[^>]*?)\sheight="[^"]*"', r"\1", text, count=1)
	return text.encode("utf-8")


def _convert_vector_image_to_svg(
	img_bytes: bytes, ext: str
) -> Tuple[Optional[bytes], Optional[Tuple[int, int]]]:
	"""Returns (svg_bytes, natural_size_px). natural_size_px is the drawing's own
	intrinsic pixel size (detected from its cropped bounding box), used as a
	sizing fallback when Word doesn't declare a placement size for it."""
	if not _SOFFICE_BIN:
		return None, None
	with tempfile.TemporaryDirectory(prefix="docximg_") as td:
		src_path = os.path.join(td, f"image.{ext}")
		with open(src_path, "wb") as fh:
			fh.write(img_bytes)
		profile_dir = os.path.join(td, f"lo_profile_{uuid.uuid4().hex}")

		def _convert(target_fmt: str) -> Optional[bytes]:
			try:
				subprocess.run(
					[
						_SOFFICE_BIN,
						f"-env:UserInstallation=file://{profile_dir}",
						"--headless",
						"--norestore",
						"--convert-to",
						target_fmt,
						"--outdir",
						td,
						src_path,
					],
					check=True,
					capture_output=True,
					timeout=25,
				)
			except Exception:
				return None
			out_path = os.path.join(td, f"image.{target_fmt}")
			if not os.path.isfile(out_path):
				return None
			with open(out_path, "rb") as fh:
				return fh.read() or None

		svg_bytes = _convert("svg")
		if not svg_bytes:
			return None, None

		# LibreOffice always renders EMF/WMF onto a full page canvas (e.g. A4)
		# regardless of how small the actual drawing is, so the artwork ends up
		# as a speck inside a mostly-blank SVG. Render the same source to PNG
		# purely to detect the drawing's real bounding box, then crop the SVG's
		# viewBox to it - keeps the output fully vector, just windowed tighter.
		natural_size: Optional[Tuple[int, int]] = None
		png_bytes = _convert("png")
		if png_bytes:
			cropped = _autocrop_bbox(png_bytes)
			if cropped:
				(x0, y0, x1, y1), (page_w, page_h) = cropped
				natural_size = (x1 - x0, y1 - y0)
				frac_bbox = (x0 / page_w, y0 / page_h, x1 / page_w, y1 / page_h)
				svg_bytes = _crop_svg_viewbox(svg_bytes, frac_bbox)

		return svg_bytes, natural_size


def _encode_image_bytes(
	img_bytes: bytes, ext: str
) -> Tuple[str, str, Optional[Tuple[int, int]]]:
	"""Returns (mime, base64, natural_size_px) for an embedded image, converting
	EMF/WMF to SVG when possible. natural_size_px is the image's own intrinsic
	pixel size (read from the file itself), used as a sizing fallback when Word
	doesn't declare a placement size (wp:extent / VML style) for this drawing."""
	ext = (ext or "").lower().lstrip(".")

	if ext in _VECTOR_EXT:
		svg_bytes, natural_size = _convert_vector_image_to_svg(img_bytes, ext)
		if svg_bytes:
			return "image/svg+xml", base64.b64encode(svg_bytes).decode("ascii"), natural_size
		# LibreOffice unavailable/failed: fall back to the raw metafile so no
		# data is lost, even though most browsers won't render it inline.
		fallback_mime = "image/x-emf" if ext == "emf" else "image/x-wmf"
		return fallback_mime, base64.b64encode(img_bytes).decode("ascii"), None

	mime = _RASTER_EXT_MIME.get(ext) or mimetypes.types_map.get(f".{ext}", "application/octet-stream")
	natural_size: Optional[Tuple[int, int]] = None
	try:
		with Image.open(io.BytesIO(img_bytes)) as im:
			natural_size = im.size
	except Exception:
		natural_size = None
	return mime, base64.b64encode(img_bytes).decode("ascii"), natural_size


def _build_image_by_rid(z: "zipfile.ZipFile") -> Dict[str, Tuple[str, str, Optional[Tuple[int, int]]]]:
	"""Maps a DOCX relationship id -> (mime, base64, natural_size_px) for every embedded image."""
	image_by_rid: Dict[str, Tuple[str, str, Optional[Tuple[int, int]]]] = {}
	try:
		rels_xml = z.read("word/_rels/document.xml.rels")
		rels_root = fromstring(rels_xml)
		for rel in rels_root.findall(".//{*}Relationship"):
			rid = rel.get("Id")
			target = rel.get("Target") or ""
			rtype = rel.get("Type") or ""
			if not rid or not target:
				continue
			if "relationships/image" not in rtype:
				continue
			norm = target.lstrip("/")
			if norm.startswith("../"):
				norm = norm.replace("../", "")
			if not norm.startswith("word/"):
				norm = f"word/{norm}"
			try:
				img_bytes = z.read(norm)
			except Exception:
				continue
			ext = (norm.rsplit(".", 1)[-1] or "").lower()
			image_by_rid[rid] = _encode_image_bytes(img_bytes, ext)
	except Exception:
		return {}
	return image_by_rid


# Word lets an image be inserted at any size, but rendering it at
# "max-width:100%" blows it up to the full width of whatever container holds
# it (an answer cell, a modal, ...), which in turn inflates the surrounding
# table row. Cap the rendered size and, where Word tells us the size the
# author actually placed the image at (wp:extent / VML style), honor that
# instead of always stretching to fill the container.
MAX_IMG_WIDTH_PX = 220
MAX_IMG_HEIGHT_PX = 220
_EMU_PER_PX = 9525  # 914400 EMU per inch / 96 px per inch


def _emu_to_px(value) -> Optional[int]:
	try:
		px = round(int(value) / _EMU_PER_PX)
		return px if px > 0 else None
	except Exception:
		return None


def _pt_to_px(value: float) -> int:
	return max(1, round(value * 96 / 72))


def _drawing_extent_px(node: etree._Element) -> Optional[Tuple[int, int]]:
	"""Looks up the ancestor w:drawing's declared width/height (wp:extent, in EMUs)."""
	anc = node.getparent()
	while anc is not None and etree.QName(anc).localname != "drawing":
		anc = anc.getparent()
	if anc is None:
		return None
	for child in anc.iter():
		if etree.QName(child).localname == "extent":
			w, h = _emu_to_px(child.get("cx")), _emu_to_px(child.get("cy"))
			if w and h:
				return w, h
	return None


def _vml_shape_size_px(node: etree._Element) -> Optional[Tuple[int, int]]:
	"""Looks up the parent v:shape's inline `style="width:..pt;height:..pt"`."""
	shape = node.getparent()
	style = (shape.get("style") if shape is not None else None) or ""
	w_match = re.search(r"width:\s*([\d.]+)pt", style)
	h_match = re.search(r"height:\s*([\d.]+)pt", style)
	if w_match and h_match:
		return _pt_to_px(float(w_match.group(1))), _pt_to_px(float(h_match.group(1)))
	return None


def _fit_within(w: int, h: int, max_w: int, max_h: int) -> Tuple[int, int]:
	scale = min(1.0, max_w / w, max_h / h)
	return max(1, round(w * scale)), max(1, round(h * scale))


def _img_tag(mime: str, b64: str, size_px: Optional[Tuple[int, int]]) -> str:
	if size_px:
		w, h = _fit_within(size_px[0], size_px[1], MAX_IMG_WIDTH_PX, MAX_IMG_HEIGHT_PX)
		size_style = f"width:{w}px;height:{h}px;max-width:100%;"
	else:
		# `min()` keeps both caps active at once. Declaring `max-width` twice in
		# one style attribute (e.g. "max-width:420px;...;max-width:100%") isn't
		# additive - the later declaration silently wins and cancels the first,
		# which let extent-less images grow to the full container width while
		# only their height stayed capped.
		size_style = f"max-width:min({MAX_IMG_WIDTH_PX}px, 100%) !important;max-height:{MAX_IMG_HEIGHT_PX}px !important;"
	return (
		f'<img src="data:{mime};base64,{b64}" alt="image" '
		f'style="{size_style}object-fit:contain;display:block;margin:0.5em 0;" />'
	)


def _is_filled(text_or_html: str) -> bool:
	s = str(text_or_html or "")
	if not s.strip():
		return False
	# Images are meaningful content in our DOCX HTML output.
	# Without this, cells that contain only <img> tags are treated as empty.
	if re.search(r"<\s*img\b", s, flags=re.I):
		return True
	# Latex marker or non-empty text
	if "$" in s:
		return True
	plain = _strip_tags(s)
	plain = re.sub(r"\s+", "", plain)
	return len(plain) > 0


def _omml_operator_to_latex(op: str) -> str:
	op = (op or "").strip()
	return {
		"∑": "\\sum",
		"∏": "\\prod",
		"∫": "\\int",
		"∮": "\\oint",
		"⋂": "\\bigcap",
		"⋃": "\\bigcup",
		"⋁": "\\bigvee",
		"⋀": "\\bigwedge",
	}.get(op, op)


def _first_child(el: etree._Element, local: str) -> Optional[etree._Element]:
	for c in el:
		if etree.QName(c).localname == local:
			return c
	return None


def _children(el: etree._Element, local: str) -> List[etree._Element]:
	return [c for c in el if etree.QName(c).localname == local]


def _clean_latex(s: str) -> str:
	s = (s or "").replace("\u00a0", " ")
	s = re.sub(r"\s+", " ", s).strip()
	# Common broken artifact from partial OMML nodes
	s = s.replace("\\frac{}{}", "")
	return s


def omml_to_latex(el: etree._Element) -> str:
	"""Best-effort OMML -> LaTeX.

	Covers the most common Word Equation constructs (fractions, scripts, radicals, n-ary, matrices).
	Unknown nodes fall back to concatenating children.
	"""
	name = etree.QName(el).localname

	# Text inside OMML runs
	if name == "t":
		return el.text or ""

	# Common wrappers
	if name in {"oMath", "oMathPara", "e", "r"}:
		return "".join(omml_to_latex(c) for c in el)

	if name == "f":
		num_wrap = _first_child(el, "num")
		den_wrap = _first_child(el, "den")
		num = _first_child(num_wrap if num_wrap is not None else el, "e")
		den = _first_child(den_wrap if den_wrap is not None else el, "e")
		num_latex = _clean_latex(omml_to_latex(num) if num is not None else "")
		den_latex = _clean_latex(omml_to_latex(den) if den is not None else "")
		if not num_latex and not den_latex:
			return ""
		# If one side is missing, keep MathJax-friendly tiny spacing placeholders.
		if not num_latex:
			num_latex = "\\,"
		if not den_latex:
			den_latex = "\\,"
		return f"\\frac{{{num_latex}}}{{{den_latex}}}"

	if name == "sSup":
		base = _first_child(el, "e")
		sup = _first_child(el, "sup")
		return f"{omml_to_latex(base) if base is not None else ''}^{{{omml_to_latex(sup) if sup is not None else ''}}}"

	if name == "sSub":
		base = _first_child(el, "e")
		sub = _first_child(el, "sub")
		return f"{omml_to_latex(base) if base is not None else ''}_{{{omml_to_latex(sub) if sub is not None else ''}}}"

	if name == "sSubSup":
		base = _first_child(el, "e")
		sub = _first_child(el, "sub")
		sup = _first_child(el, "sup")
		return (
			f"{omml_to_latex(base) if base is not None else ''}"
			f"_{{{omml_to_latex(sub) if sub is not None else ''}}}"
			f"^{{{omml_to_latex(sup) if sup is not None else ''}}}"
		)

	if name == "rad":
		e = _first_child(el, "e")
		deg = _first_child(el, "deg")
		deg_e = _first_child(deg, "e") if deg is not None else None
		if deg_e is not None and omml_to_latex(deg_e).strip():
			return f"\\sqrt[{omml_to_latex(deg_e)}]{{{omml_to_latex(e) if e is not None else ''}}}"
		return f"\\sqrt{{{omml_to_latex(e) if e is not None else ''}}}"

	if name == "nary":
		pr = _first_child(el, "naryPr")
		chr_el = pr.find("m:chr", namespaces=NS) if pr is not None else None
		chr_val = None
		if chr_el is not None:
			chr_val = chr_el.get(f"{{{W_NS}}}val") or chr_el.get("val")
		op = _omml_operator_to_latex(chr_val or "")

		sub = _first_child(el, "sub")
		sup = _first_child(el, "sup")
		e = _first_child(el, "e")

		out = op or "\\sum"
		if sub is not None and omml_to_latex(sub).strip():
			out += f"_{{{omml_to_latex(sub)}}}"
		if sup is not None and omml_to_latex(sup).strip():
			out += f"^{{{omml_to_latex(sup)}}}"
		body = omml_to_latex(e) if e is not None else ""
		if body:
			out += f"{{{body}}}"
		return out

	if name == "limLow":
		e = _first_child(el, "e")
		lim = _first_child(el, "lim")
		lim_e = _first_child(lim, "e") if lim is not None else None
		lim_s = omml_to_latex(lim_e) if lim_e is not None else ""
		return f"\\lim_{{{lim_s}}}{omml_to_latex(e) if e is not None else ''}"

	if name == "limUpp":
		e = _first_child(el, "e")
		lim = _first_child(el, "lim")
		lim_e = _first_child(lim, "e") if lim is not None else None
		lim_s = omml_to_latex(lim_e) if lim_e is not None else ""
		return f"\\lim^{{{lim_s}}}{omml_to_latex(e) if e is not None else ''}"

	if name == "d":
		# Delimiters (parentheses, brackets, etc.)
		pr = _first_child(el, "dPr")
		beg = pr.find("m:begChr", namespaces=NS) if pr is not None else None
		end = pr.find("m:endChr", namespaces=NS) if pr is not None else None
		beg_val = (beg.get(f"{{{W_NS}}}val") or beg.get("val")) if beg is not None else "("
		end_val = (end.get(f"{{{W_NS}}}val") or end.get("val")) if end is not None else ")"
		inner = "".join(omml_to_latex(c) for c in _children(el, "e"))
		return f"\\left{beg_val}{inner}\\right{end_val}"

	if name == "bar":
		e = _first_child(el, "e")
		return f"\\overline{{{omml_to_latex(e) if e is not None else ''}}}"

	if name == "m":
		# matrix: m:mr rows, each has m:e entries
		rows = []
		for mr in _children(el, "mr"):
			cells = [omml_to_latex(e) for e in _children(mr, "e")]
			rows.append(" & ".join(cells))
		body = " \\\\ ".join(rows)
		return f"\\begin{{matrix}}{body}\\end{{matrix}}"

	# Fallback: concat children
	return "".join(omml_to_latex(c) for c in el)


@dataclass
class ParsedQuestion:
	question: str
	options: List[str]
	correct_answer_index: int
	points: int


def _walk_node_to_html(tc: etree._Element, image_by_rid: Optional[dict] = None) -> str:
	"""Extract cell contents (text + equations) into safe inline HTML.

	We output escaped text + `<br/>` + `$...$` LaTeX blocks. Frontend later turns LaTeX into SVG.
	"""
	parts: List[str] = []
	seen_image_rids = set()

	def walk(node: etree._Element):
		ns = etree.QName(node).namespace
		ln = etree.QName(node).localname

		# Embedded images inside Word tables (w:drawing -> a:blip r:embed="rIdX")
		if image_by_rid and ln == "blip":
			rid = node.get(f"{{{R_NS}}}embed")
			if rid and rid in image_by_rid:
				if rid in seen_image_rids:
					return
				seen_image_rids.add(rid)
				mime, b64, natural_size = image_by_rid[rid]
				# Prefer the size Word says the author placed the drawing at; if
				# that's missing, fall back to the image's own intrinsic size
				# rather than an arbitrary generic box.
				parts.append(_img_tag(mime, b64, _drawing_extent_px(node) or natural_size))
				return

		# Older Word exports can embed images via VML:
		#   <w:pict><v:shape>...<v:imagedata r:id="rIdX"/></v:shape></w:pict>
		if image_by_rid and ln == "imagedata":
			rid = node.get(f"{{{R_NS}}}id") or node.get(f"{{{R_NS}}}embed")
			if rid and rid in image_by_rid:
				if rid in seen_image_rids:
					return
				seen_image_rids.add(rid)
				mime, b64, natural_size = image_by_rid[rid]
				parts.append(_img_tag(mime, b64, _vml_shape_size_px(node) or natural_size))
				return

		# Word line breaks
		if ns == W_NS and ln in {"br", "cr"}:
			parts.append("<br/>")
			return

		# Runs formatted with Word's Subscript/Superscript button (w:rPr/w:vertAlign)
		# rather than the Equation Editor - common for chemical formulas like C2H5COOH
		# typed as plain text with the "2" set to subscript. Without this, the
		# vertAlign is silently dropped and the formula collapses to flat text.
		if ns == W_NS and ln == "r":
			rpr = _first_child(node, "rPr")
			vert_align = None
			if rpr is not None:
				va = _first_child(rpr, "vertAlign")
				if va is not None:
					vert_align = va.get(f"{{{W_NS}}}val") or va.get("val")
			if vert_align in ("subscript", "superscript"):
				start = len(parts)
				for child in list(node):
					walk(child)
				inner = "".join(parts[start:])
				del parts[start:]
				if inner:
					tag = "sub" if vert_align == "subscript" else "sup"
					parts.append(f"<{tag}>{inner}</{tag}>")
				return

		# Word text
		if ns == W_NS and ln == "t":
			if node.text:
				parts.append(_escape_html(node.text))
			return

		# Word equations (OMML). oMathPara is Word's own "display equation on its
		# own line" wrapper (typically a bigger fraction/sum/matrix meant to be
		# shown large); a bare oMath sits inline inside running text. Emitting
		# both as inline `$...$` forces the frontend to squeeze a potentially
		# tall equation into a single text line's height, which is what mangles
		# them - so keep the distinction and use `$$...$$` for oMathPara.
		if ns == M_NS and ln in {"oMath", "oMathPara"}:
			latex = omml_to_latex(node)
			latex = (latex or "").strip()
			latex = latex.replace("$", "")
			if latex:
				parts.append(f"$${latex}$$" if ln == "oMathPara" else f"${latex}$")
			return

		for child in list(node):
			walk(child)

	walk(tc)

	out = "".join(parts)
	out = re.sub(r"(<br\s*/?>\s*){3,}", "<br/><br/>", out, flags=re.I)
	return out.strip()


def _tc_colspan(tc: etree._Element) -> int:
	grid = tc.find("w:tcPr/w:gridSpan", namespaces=NS)
	if grid is None:
		return 1
	val = grid.get(f"{{{W_NS}}}val") or grid.get("val")
	try:
		return max(1, int(val or "1"))
	except Exception:
		return 1


def _extract_docx_tables(document_xml: bytes) -> List[etree._Element]:
	root = fromstring(document_xml)
	# Word doc body tables
	return root.findall(".//w:tbl", namespaces=NS)


def _extract_docx_table_rows(tbl: etree._Element, image_by_rid: Optional[dict] = None) -> List[List[str]]:
	rows: List[List[str]] = []
	prev_row: List[str] = []
	for tr in tbl.findall("w:tr", namespaces=NS):
		out_row: List[str] = []
		col = 0
		for tc in tr.findall("w:tc", namespaces=NS):
			html = _walk_node_to_html(tc, image_by_rid=image_by_rid)

			# vertical merge support (simple carry-down)
			vmerge = tc.find("w:tcPr/w:vMerge", namespaces=NS)
			if vmerge is not None:
				val = vmerge.get(f"{{{W_NS}}}val") or vmerge.get("val") or "continue"
				if val != "restart" and col < len(prev_row) and prev_row[col]:
					html = prev_row[col]

			span = _tc_colspan(tc)
			# Ensure list length
			while len(out_row) <= col:
				out_row.append("")
			out_row[col] = html
			for s in range(1, span):
				while len(out_row) <= col + s:
					out_row.append("")
				if not out_row[col + s]:
					out_row[col + s] = ""
			col += span

		rows.append(out_row)
		prev_row = out_row
	return rows


def _score_table(rows: List[List[str]]) -> int:
	row_count = len(rows)
	max_cols = max((len(r) for r in rows), default=0)
	has_template_cols = 1 if max_cols >= 7 else 0
	return has_template_cols * 1_000_000 + max_cols * 1_000 + row_count


def parse_docx_questions(file_bytes: bytes) -> Tuple[List[ParsedQuestion], List[str]]:
	"""Parse DOCX template table into questions.

	Output question/option strings as HTML that may include `$...$` LaTeX markers.
	"""
	errors: List[str] = []
	questions: List[ParsedQuestion] = []
	lenient_correct = _env_truthy("IMPORT_LENIENT_CORRECT_ANSWER", default=False)

	with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
		xml = z.read("word/document.xml")
		image_by_rid = _build_image_by_rid(z)

	tables = _extract_docx_tables(xml)
	if not tables:
		return [], ["DOCX ichida jadval topilmadi"]

	best_rows: Optional[List[List[str]]] = None
	best_score = -1
	for t in tables:
		rows = _extract_docx_table_rows(t, image_by_rid=image_by_rid)
		score = _score_table(rows)
		if score > best_score:
			best_score = score
			best_rows = rows

	rows = best_rows or []
	if not rows:
		return [], ["DOCX jadvalini o‘qib bo‘lmadi"]

	def _row_texts(row: List[str]) -> List[str]:
		return [_strip_tags(c or "").lower() for c in (row or [])]

	def _row_join(row: List[str]) -> str:
		return " ".join(_row_texts(row))

	def headerish(row: List[str]) -> bool:
		# Some templates have merged cells or shifted columns; be tolerant.
		texts = _row_texts(row)
		joined = " ".join(texts)
		has_savol = any("savol" in t for t in texts)
		has_togri = ("to'g'ri" in joined) or ("toʻgʻri" in joined) or ("tog'ri" in joined)
		has_ball = "ball" in joined
		has_variants = any(re.search(r"\b[abcd]\)?\b", t) for t in texts) or "a)" in joined or "b)" in joined
		return has_savol and (has_togri or has_ball) and (has_variants or has_ball)

	def titleish(row: List[str]) -> bool:
		# Title row often spans all columns and contains words like "shablon"/"import".
		joined = _row_join(row)
		filled_cells = sum(1 for t in _row_texts(row) if t.strip())
		return filled_cells <= 2 and ("shablon" in joined or "import" in joined)

	# Skip title/header rows (anything before the header row, plus header row itself)
	header_idx = next((i for i, r in enumerate(rows) if len(r) >= 5 and headerish(r)), -1)
	start_idx = header_idx + 1 if header_idx >= 0 else 0

	for i, row in enumerate(rows[start_idx:], start=start_idx):
		# Need at least 7 columns
		if len(row) < 7:
			continue
		if headerish(row) or titleish(row):
			continue

		q_html = row[0]
		opts_html = [row[1], row[2], row[3], row[4]]
		correct_raw = _strip_tags(row[5]).strip().upper()
		# Normalize common Cyrillic lookalikes users may paste from Word.
		correct_raw = correct_raw.translate(str.maketrans({
			"А": "A",
			"В": "B",
			"С": "C",
			"Д": "D",
		}))
		# Accept inputs like "A)", "A.", "javob: B", etc.
		m = re.search(r"\b([ABCD])\b", correct_raw) or re.search(r"([ABCD])", correct_raw)
		if m:
			correct_letter = m.group(1)
		else:
			# Also accept 1-4 as A-D
			m_num = re.search(r"\b([1-4])\b", correct_raw)
			if m_num:
				correct_letter = {"1": "A", "2": "B", "3": "C", "4": "D"}[m_num.group(1)]
			else:
				correct_letter = ""
		points_raw = _strip_tags(row[6]).strip() or "1"
		try:
			points = int(float(points_raw))
		except Exception:
			points = 1

		if not _is_filled(q_html):
			errors.append(
				f"Qator {i + 1}: Savol matni bo‘sh (matn/rasm/formula topilmadi)"
			)
			continue
		if points < 1 or points > 10:
			errors.append(f"Qator {i + 1}: Ball 1-10 oralig‘ida bo‘lishi kerak")
			continue

		entries = []
		for letter, html in zip(["A", "B", "C", "D"], opts_html):
			if _is_filled(html):
				entries.append((letter, html))

		filled_letters = [l for (l, _) in entries]
		correct_letter_display = correct_letter or (correct_raw.strip() if correct_raw else "")

		if len(entries) < 2:
			errors.append(
				f"Qator {i + 1}: Kamida 2 ta variant bo‘lishi kerak. "
				f"To‘ldirilgan variantlar: {', '.join(filled_letters) if filled_letters else 'yo‘q'}."
			)
			continue

		idx = next((j for j, (letter, _) in enumerate(entries) if letter == correct_letter), -1)
		if idx < 0:
			if lenient_correct and correct_letter in {"A", "B", "C", "D"} and len(entries) > 0:
				# Map A/B/C/D to the 1st/2nd/3rd/4th FILLED option.
				# Example: filled = [A, C, D], correct=B -> choose 2nd filled => C.
				desired_pos = {"A": 0, "B": 1, "C": 2, "D": 3}[correct_letter]
				chosen = min(desired_pos, len(entries) - 1)
				errors.append(
					f"Qator {i + 1}: To‘g‘ri javob '{correct_letter_display}' bo‘lib, lekin o‘sha ustun bo‘sh. "
					f"Lenient rejim: {correct_letter} -> {filled_letters[chosen]} (to‘ldirilgan variantlar: {', '.join(filled_letters)})."
				)
				idx = chosen
			else:
				errors.append(
					f"Qator {i + 1}: To‘g‘ri javob A/B/C/D bo‘lishi va o‘sha variant to‘ldirilgan bo‘lishi kerak. "
					f"Keltirilgan: {correct_letter_display or 'bo‘sh'}. "
					f"To‘ldirilgan variantlar: {', '.join(filled_letters) if filled_letters else 'yo‘q'}."
				)
				continue

		questions.append(
			ParsedQuestion(
				question=q_html,
				options=[v for _, v in entries],
				correct_answer_index=idx,
				points=points,
			)
		)

	return questions, errors


WORD_QUESTION_OPTION_RE = re.compile(r"^(\+)?\s*[A-Da-d]\)\s*")


def _extract_docx_body_paragraphs(document_xml: bytes) -> List[etree._Element]:
	root = fromstring(document_xml)
	body = root.find("w:body", namespaces=NS)
	if body is None:
		return []
	return [c for c in body if etree.QName(c).localname == "p"]


def parse_docx_word_questions(file_bytes: bytes) -> Tuple[List[dict], List[str]]:
	"""Parse a DOCX using the plain-text convention:

		#Savol matni
		A) variant
		+C) to'g'ri variant
		D) variant

	- A question starts on a paragraph beginning with '#'.
	- Each option paragraph starts with a letter A-D followed by ')'.
	- The correct option's letter is prefixed with '+' (e.g. "+C)").
	- Anything else (extra paragraphs, embedded images) is attached to whatever
	  is currently open: the question text if no option has started yet,
	  otherwise the most recently seen option.

	Reuses the same lxml-based HTML walker as the table importer, so embedded
	images (converted EMF/WMF -> SVG where possible) and Word equations (OMML
	-> `$...$` LaTeX) are handled identically here.
	"""
	questions: List[dict] = []
	errors: List[str] = []
	current: Optional[dict] = None
	q_number = 0

	def flush():
		nonlocal current, q_number
		if current is None:
			return
		q_number += 1
		draft = current
		current = None

		html = draft["html"].strip()
		plain = _strip_tags(html)
		has_image = bool(re.search(r"<\s*img\b", html, flags=re.I))
		if not plain and not has_image:
			errors.append(f"{q_number}-savol: matni bo'sh")
			return

		if len(draft["options"]) < 2:
			errors.append(f"{q_number}-savol: kamida 2 ta variant bo'lishi kerak")
			return

		correct_index = next((i for i, o in enumerate(draft["options"]) if o["correct"]), -1)
		if correct_index < 0:
			errors.append(f"{q_number}-savol: to'g'ri variant ('+') belgilanmagan")
			return

		questions.append(
			{
				"text": html,
				"type": "multiple_choice",
				"points": 1,
				"answers": [
					{"text": opt["html"].strip(), "isCorrect": i == correct_index, "order": i}
					for i, opt in enumerate(draft["options"])
				],
			}
		)

	with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
		xml = z.read("word/document.xml")
		image_by_rid = _build_image_by_rid(z)

	paragraphs = _extract_docx_body_paragraphs(xml)

	for p in paragraphs:
		html = _walk_node_to_html(p, image_by_rid=image_by_rid)
		raw_text = "".join(p.itertext()).strip()

		if not html and not raw_text:
			continue

		if raw_text.startswith("#"):
			flush()
			stripped_html = re.sub(r"^#\s*", "", html, count=1).strip()
			current = {"html": stripped_html, "options": []}
			continue

		match = WORD_QUESTION_OPTION_RE.match(raw_text)
		if match and current is not None:
			is_correct = raw_text.lstrip().startswith("+")
			stripped_html = WORD_QUESTION_OPTION_RE.sub("", html, count=1).strip()
			current["options"].append({"html": stripped_html, "correct": is_correct})
			continue

		# Stray content (e.g. an image on its own paragraph) belongs to whatever
		# is currently open: the question text, or the most recent option.
		if current is not None and html:
			if not current["options"]:
				current["html"] = f'{current["html"]}<br/>{html}' if current["html"] else html
			else:
				last = current["options"][-1]
				last["html"] = f'{last["html"]}<br/>{html}' if last["html"] else html

	flush()
	return questions, errors


def parse_xlsx_questions(file_bytes: bytes) -> Tuple[List[ParsedQuestion], List[str]]:
	errors: List[str] = []
	questions: List[ParsedQuestion] = []

	wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=False)
	ws = wb.worksheets[0]

	# Expect header in first row; start from row 2
	for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
		values = ["" if v is None else str(v) for v in row]
		# ensure at least 7
		if len(values) < 7:
			continue
		q = values[0].strip()
		a, b, c, d = (values[1].strip(), values[2].strip(), values[3].strip(), values[4].strip())
		correct = values[5].strip().upper()
		points_raw = values[6].strip() or "1"
		try:
			points = int(float(points_raw))
		except Exception:
			points = 1

		if not q:
			errors.append(f"Qator {i}: Savol matni bo‘sh")
			continue
		if points < 1 or points > 10:
			errors.append(f"Qator {i}: Ball 1-10 oralig‘ida bo‘lishi kerak")
			continue

		entries = [("A", a), ("B", b), ("C", c), ("D", d)]
		entries = [(l, v) for (l, v) in entries if v]
		if len(entries) < 2:
			errors.append(f"Qator {i}: Kamida 2 ta variant bo‘lishi kerak")
			continue

		idx = next((j for j, (letter, _) in enumerate(entries) if letter == correct), -1)
		if idx < 0:
			errors.append(
				f"Qator {i}: To‘g‘ri javob A/B/C/D bo‘lishi va o‘sha variant to‘ldirilgan bo‘lishi kerak"
			)
			continue

		questions.append(
			ParsedQuestion(
				question=_escape_html(q),
				options=[_escape_html(v) for _, v in entries],
				correct_answer_index=idx,
				points=points,
			)
		)

	return questions, errors
